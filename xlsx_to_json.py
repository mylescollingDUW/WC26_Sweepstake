"""Convert Tracker.xlsx -> tracker.json (the dashboard's source of truth).

Run this from the project folder whenever you've made bulk edits in
Excel and want them reflected on the dashboard:

    python xlsx_to_json.py

The script reads the Tracker.xlsx sitting next to it and overwrites
tracker.json in the same folder.
"""
import json
import datetime
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
SRC = ROOT / "Tracker.xlsx"
OUT = ROOT / "tracker.json"

AWARD_KEY_BY_LABEL = {
    "Tournament Winner":        "winner",
    "2nd Place":                "runnerup",
    "3rd Place":                "third",
    "Golden Boot (Top Scorer)": "goldenBoot",
}


def cell(ws, row, col, default=None):
    v = ws.cell(row=row, column=col).value
    return default if v is None or (isinstance(v, str) and not v.strip()) else v


def s(v, default=""):
    return "" if v is None else str(v).strip() or default


def num(v, default=0):
    if v is None or v == "":
        return default
    try:
        return float(v) if isinstance(v, float) else int(v)
    except (ValueError, TypeError):
        return default


def numornull(v):
    if v is None or v == "":
        return None
    try:
        n = float(v)
        return int(n) if n.is_integer() else n
    except (ValueError, TypeError):
        return None


def date_str(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    return str(v)


wb = load_workbook(SRC, data_only=True)  # data_only: read formula results

# Participants -------------------------------------------------------
ws = wb["Participants"]
teams = []
for r in range(2, ws.max_row + 1):
    team = s(cell(ws, r, 2))
    if not team:
        continue
    teams.append({
        "group": s(cell(ws, r, 1)),
        "team": team,
        "entrant": s(cell(ws, r, 3)),
    })

# Match Data ---------------------------------------------------------
ws = wb["Match Data"]
# header row 1; columns mapped by name to be resilient
headers = [s(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
col = {h: i + 1 for i, h in enumerate(headers) if h}

def mc(r, name, default=0):
    return num(cell(ws, r, col[name]), default) if name in col else default
def mc_n(r, name):
    return numornull(cell(ws, r, col[name])) if name in col else None
def mc_s(r, name):
    return s(cell(ws, r, col[name])) if name in col else ""

matches = []
mid = 0
for r in range(2, ws.max_row + 1):
    home = mc_s(r, "Home Team")
    away = mc_s(r, "Away Team")
    if not home or not away:
        continue
    mid += 1
    matches.append({
        "id": f"M{mid:03d}",
        "date": date_str(cell(ws, r, col.get("Date", 0))) if "Date" in col else None,
        "stage": mc_s(r, "Stage") or "Group",
        "homeTeam": home,
        "awayTeam": away,
        "homeScore": mc_n(r, "Home Score"),
        "awayScore": mc_n(r, "Away Score"),
        "minutes": mc(r, "Minutes", 90) or 90,
        "homeShots":      mc(r, "Home Shots"),
        "awayShots":      mc(r, "Away Shots"),
        "homeSoT":        mc(r, "Home SoT"),
        "awaySoT":        mc(r, "Away SoT"),
        "homePossession": mc(r, "Home Possession %"),
        "awayPossession": mc(r, "Away Possession %"),
        "homeFouls":      mc(r, "Home Fouls"),
        "awayFouls":      mc(r, "Away Fouls"),
        "homeYellow":     mc(r, "Home Yellow"),
        "awayYellow":     mc(r, "Away Yellow"),
        "homeRed":        mc(r, "Home Red"),
        "awayRed":        mc(r, "Away Red"),
        "homeOffsides":   mc(r, "Home Offsides"),
        "awayOffsides":   mc(r, "Away Offsides"),
        "homeCorners":    mc(r, "Home Corners"),
        "awayCorners":    mc(r, "Away Corners"),
    })

# Awards -------------------------------------------------------------
ws = wb["Awards"]
awards = {}
for r in range(2, ws.max_row + 1):
    label = s(cell(ws, r, 1))
    if not label:
        continue
    key = AWARD_KEY_BY_LABEL.get(label)
    if not key:
        continue
    awards[key] = {
        "player": s(cell(ws, r, 2)),
        "country": s(cell(ws, r, 3)),
    }
# Ensure all 4 keys exist even if blank
for k in AWARD_KEY_BY_LABEL.values():
    awards.setdefault(k, {"player": "", "country": ""})

# Golden Boot Tracker ------------------------------------------------
ws = wb["Golden Boot Tracker"]
gb = []
for r in range(2, ws.max_row + 1):
    player = s(cell(ws, r, 1))
    if not player:
        continue
    gb.append({
        "player": player,
        "country": s(cell(ws, r, 2)),
        "goals": num(cell(ws, r, 3)),
    })

doc = {
    "version": 1,
    "teams": teams,
    "matches": matches,
    "awards": awards,
    "goldenBoot": gb,
}

OUT.write_text(json.dumps(doc, indent=2, ensure_ascii=False), encoding="utf-8")
print(f"Wrote {OUT}")
print(f"  teams:      {len(teams)}")
print(f"  matches:    {len(matches)}")
print(f"  awards:     {sum(1 for v in awards.values() if v['country'])} filled")
print(f"  goldenBoot: {len(gb)}")
